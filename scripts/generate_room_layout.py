#!/usr/bin/env python3
"""Generate room layout assets from the source Excel workbook.

The workbook is treated as the single source of truth. This script reads the
sheet layout, infers rooms, buildings, floors, room types, capacities, and
corridor relationships, then writes normalized assets under ``assets/``.

Usage:
    python3 scripts/generate_room_layout.py --input /path/to/숙소배정표.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOM_RE = re.compile(r"^(\d{3})호(?:\(([^)]+)\))?$")
SECTION_RE = re.compile(r"(?P<building>.+?)(?P<floor>\d)층$")


@dataclass(frozen=True)
class Section:
    building: str
    floor: int
    start_row: int
    end_row: int
    corridor_row: int | None


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def classify_room_type(raw_suffix: str | None, floor_default: str | None = None) -> tuple[str, int, str]:
    """Return (room_type, capacity, source_kind)."""
    if raw_suffix:
        suffix = normalize_text(raw_suffix)
        bed_match = re.search(r"침대(\d+)", suffix)
        if bed_match:
            capacity = int(bed_match.group(1))
            if capacity == 1:
                return "single", 1, "explicit"
            if capacity == 2:
                return "twin", 2, "explicit"
            return f"bed_{capacity}", capacity, "explicit"

        person_match = re.search(r"(\d+)인실", suffix)
        if person_match:
            capacity = int(person_match.group(1))
            return f"{capacity}_person", capacity, "explicit"

    if floor_default:
        if floor_default == "single":
            return "single", 1, "inferred"
        if floor_default == "twin":
            return "twin", 2, "inferred"
        if floor_default == "ondol_4":
            return "ondol_4", 4, "inferred"
        if floor_default.endswith("_person"):
            capacity = int(floor_default.split("_", 1)[0])
            return floor_default, capacity, "inferred"

    return "unknown", 0, "inferred"


def room_type_label(room_type: str, capacity: int) -> str:
    if room_type == "unavailable":
        return "사용불가"
    if room_type == "single":
        return "1인실"
    if room_type == "twin":
        return "2인실"
    if room_type == "ondol_4":
        return "4인실 온돌"
    if room_type.endswith("_person"):
        return f"{capacity}인실"
    if room_type.startswith("bed_"):
        return f"침대{capacity}"
    return "미정"


def room_type_family(room_type: str) -> str:
    if room_type == "unavailable":
        return "service"
    if room_type in {"single", "twin"}:
        return "standard"
    if room_type == "ondol_4":
        return "ondol"
    if room_type.endswith("_person"):
        return "capacity"
    if room_type.startswith("bed_"):
        return "bed"
    return "unknown"


def infer_default_type(room_entries: list[dict[str, Any]], section: Section) -> str:
    explicit_types = [entry["room_type"] for entry in room_entries if entry["source_kind"] == "explicit"]
    if not explicit_types:
        # Workbook-specific heuristic: 휴락동 2층/3층 unlabeled rooms are 온돌 4인실.
        if section.building == "휴락동" and section.floor in {2, 3}:
            return "ondol_4"
        return "twin"
    counts = Counter(explicit_types)
    return counts.most_common(1)[0][0]


def apply_workbook_overrides(room: dict[str, Any], section: Section) -> None:
    """Apply workbook layout rules that are implied by position rather than suffix text."""
    if section.building == "휴락동" and section.floor == 4 and room["corridor_relationship"] == "south_of_corridor":
        room["room_type"] = "ondol_4"
        room["capacity"] = 4
        room["room_type_label"] = room_type_label("ondol_4", 4)
        room["room_type_family"] = room_type_family("ondol_4")
        room["source_kind"] = "position_override"


def apply_service_room_overrides(section_rooms: list[dict[str, Any]], section_spaces: list[dict[str, Any]]) -> None:
    """Mark rooms blocked by service labels that occupy the same workbook column."""
    service_by_position = {(space["row"], space["column"]): space for space in section_spaces if space["type"] == "service_space"}
    blocking_labels = {"당직실", "비품실", "미화원실"}

    for room in section_rooms:
        blocking_space = None
        for adjacent_row in (room["row"] + 1, room["row"] - 1):
            space = service_by_position.get((adjacent_row, room["column"]))
            if space and normalize_text(space["label"]) in blocking_labels:
                blocking_space = space
                break

        if not blocking_space:
            continue

        room["room_type"] = "unavailable"
        room["capacity"] = 0
        room["room_type_label"] = str(blocking_space["label"]).strip()
        room["room_type_family"] = room_type_family("unavailable")
        room["source_kind"] = "service_override"
        room["unavailable"] = True
        room["unavailable_reason"] = str(blocking_space["label"]).strip()


def find_sections(ws) -> list[Section]:
    section_rows: list[tuple[int, str, int]] = []
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row, 2).value
        normalized = normalize_text(cell_value)
        match = SECTION_RE.match(normalized)
        if match:
            section_rows.append((row, match.group("building"), int(match.group("floor"))))

    sections: list[Section] = []
    for index, (row, building, floor) in enumerate(section_rows):
        end_row = section_rows[index + 1][0] - 1 if index + 1 < len(section_rows) else ws.max_row
        corridor_row = None
        for scan_row in range(row, end_row + 1):
            row_values = [normalize_text(ws.cell(scan_row, col).value) for col in range(1, ws.max_column + 1)]
            if any(value == "복도" for value in row_values):
                corridor_row = scan_row
                break
        sections.append(Section(building=building, floor=floor, start_row=row, end_row=end_row, corridor_row=corridor_row))
    return sections


def is_service_space(normalized_text: str) -> bool:
    if not normalized_text:
        return False
    if ROOM_RE.match(normalized_text):
        return False
    if SECTION_RE.match(normalized_text):
        return False
    service_keywords = {
        "복도",
        "계단",
        "엘리베이터",
        "ELV",
        "세탁실",
        "벨브실",
        "조리원실",
        "식당→",
        "식당",
        "공동화장실",
        "당직실",
        "비품실",
        "미화원실",
    }
    return normalized_text in service_keywords or normalized_text.endswith("→") or "실" in normalized_text


def extract_room_entries(ws, sections: list[Section]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rooms: list[dict[str, Any]] = []
    spaces: list[dict[str, Any]] = []

    for section_index, section in enumerate(sections):
        section_rooms: list[dict[str, Any]] = []
        section_spaces: list[dict[str, Any]] = []

        for row in range(section.start_row, section.end_row + 1):
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row, col).value
                normalized = normalize_text(value)
                if not normalized:
                    continue

                room_match = ROOM_RE.match(normalized)
                if room_match:
                    room_number = int(room_match.group(1))
                    suffix = room_match.group(2)
                    room_type, capacity, source_kind = classify_room_type(suffix)
                    entry = {
                        "section_index": section_index,
                        "building": section.building,
                        "floor": section.floor,
                        "room_number": room_number,
                        "room_label": f"{room_number}호",
                        "room_type": room_type,
                        "capacity": capacity,
                        "room_type_label": room_type_label(room_type, capacity),
                        "room_type_family": room_type_family(room_type),
                        "source_kind": source_kind,
                        "source_text": str(value),
                        "cell": f"{get_column_letter(col)}{row}",
                        "row": row,
                        "column": col,
                        "corridor_row": section.corridor_row,
                        "corridor_relationship": "north_of_corridor" if section.corridor_row and row < section.corridor_row else "south_of_corridor" if section.corridor_row and row > section.corridor_row else "unknown",
                        "source_sheet": ws.title,
                    }
                    section_rooms.append(entry)
                    continue

                if is_service_space(normalized):
                    space_type = "corridor" if normalized == "복도" else "service_space"
                    section_spaces.append(
                        {
                            "section_index": section_index,
                            "building": section.building,
                            "floor": section.floor,
                            "label": str(value).strip(),
                            "cell": f"{get_column_letter(col)}{row}",
                            "row": row,
                            "column": col,
                            "type": space_type,
                            "source_sheet": ws.title,
                        }
                    )

        default_type = infer_default_type(section_rooms, section)
        for entry in section_rooms:
            if entry["source_kind"] != "explicit":
                inferred_type, capacity, _ = classify_room_type(None, floor_default=default_type)
                entry["room_type"] = inferred_type
                entry["capacity"] = capacity
                entry["room_type_label"] = room_type_label(inferred_type, capacity)
                entry["room_type_family"] = room_type_family(inferred_type)
            entry["corridor_relationship"] = (
                "north_of_corridor"
                if section.corridor_row and entry["row"] < section.corridor_row
                else "south_of_corridor"
                if section.corridor_row and entry["row"] > section.corridor_row
                else "unknown"
            )
            apply_workbook_overrides(entry, section)

        apply_service_room_overrides(section_rooms, section_spaces)

        for entry in section_rooms:
            rooms.append(entry)

        spaces.extend(section_spaces)

    rooms.sort(key=lambda item: (item["section_index"], item["row"], item["column"], item["room_number"]))
    spaces.sort(key=lambda item: (item["section_index"], item["row"], item["column"], item["label"]))
    return rooms, spaces


def group_structure(rooms: list[dict[str, Any]], spaces: list[dict[str, Any]], sections: list[Section]) -> dict[str, Any]:
    buildings: list[dict[str, Any]] = []
    rooms_by_building_floor = defaultdict(list)
    spaces_by_building_floor = defaultdict(list)
    section_lookup = {(section.building, section.floor): section for section in sections}

    for room in rooms:
        rooms_by_building_floor[(room["building"], room["floor"])].append(room)
    for space in spaces:
        spaces_by_building_floor[(space["building"], space["floor"])].append(space)

    for building in dict.fromkeys(section.building for section in sections):
        floors: list[dict[str, Any]] = []
        building_sections = [section for section in sections if section.building == building]
        for section in building_sections:
            floor_rooms = rooms_by_building_floor[(section.building, section.floor)]
            floor_spaces = spaces_by_building_floor[(section.building, section.floor)]
            floors.append(
                {
                    "floor": section.floor,
                    "label": f"{section.floor}층",
                    "corridor_row": section.corridor_row,
                    "room_count": len(floor_rooms),
                    "rooms": [
                        {
                            "room_number": room["room_number"],
                            "room_label": room["room_label"],
                            "room_type": room["room_type"],
                            "room_type_label": room["room_type_label"],
                            "room_type_family": room["room_type_family"],
                            "capacity": room["capacity"],
                            "corridor_relationship": room["corridor_relationship"],
                            "cell": room["cell"],
                            "row": room["row"],
                            "column": room["column"],
                            "source_text": room["source_text"],
                            "source_kind": room["source_kind"],
                            "unavailable": room.get("unavailable", False),
                            "unavailable_reason": room.get("unavailable_reason", ""),
                        }
                        for room in floor_rooms
                    ],
                    "service_spaces": floor_spaces,
                }
            )
        buildings.append({"building": building, "floors": floors})

    return {
        "source": "숙소배정표.xlsx",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "buildings": buildings,
    }


def build_room_rules_md(structure: dict[str, Any], rooms: list[dict[str, Any]]) -> str:
    lines = [
        "# Room Layout Rules",
        "",
        "This file documents how the room layout assets were inferred from the Excel workbook.",
        "",
        "## Source of truth",
        "- Workbook: `숙소배정표.xlsx`",
        "- The parser reads the sheet text and cell positions directly.",
        "- No room numbers are hardcoded into the generator.",
        "",
        "## Inference rules",
        "1. Section headers are detected from the `B` column text such as `휴락동4층` or `동락홀2층`.",
        "2. Room cells are detected with the pattern `###호` and optional room-type suffixes in parentheses.",
        "3. Explicit suffixes override inference:",
        "   - `침대1` -> `single` -> capacity `1`",
        "   - `침대2` -> `twin` -> capacity `2`",
        "   - `N인실` -> capacity `N`",
        "4. When a room has no explicit suffix, the floor default is used.",
        "5. The floor default is the most common explicit room type found in that floor block.",
        "6. If a floor has no explicit room-type labels, workbook-specific heuristics are applied.",
        "   - `휴락동 2층` and `휴락동 3층` unlabeled rooms are interpreted as `4인실 온돌`.",
        "   - other unlabeled rooms fall back to `2인실`.",
        "7. Position-specific workbook overrides are applied after default inference.",
        "   - `휴락동 4층` rooms south of the corridor are interpreted as `4인실 온돌`.",
        "8. Rooms adjacent to same-column service labels such as `당직실`, `비품실`, or `미화원실` are marked as unavailable.",
        "9. Corridor relationship is inferred from the room row relative to the corridor row in the same floor block.",
        "",
        "## Current workbook findings",
    ]

    by_building_floor = defaultdict(list)
    for room in rooms:
        by_building_floor[(room["building"], room["floor"])].append(room)

    for building in dict.fromkeys(item["building"] for item in structure["buildings"]):
        lines.append(f"### {building}")
        building_data = next(item for item in structure["buildings"] if item["building"] == building)
        for floor in building_data["floors"]:
            floor_rooms = by_building_floor[(building, floor["floor"])]
            type_counts = Counter(room["room_type_label"] for room in floor_rooms)
            label_summary = ", ".join(f"{label} {count}개" for label, count in sorted(type_counts.items()))
            corridor = floor["corridor_row"] if floor["corridor_row"] is not None else "unknown"
            lines.append(f"- {floor['label']}: {floor['room_count']}개 방, corridor row {corridor}, room types: {label_summary}")
        lines.append("")

    lines.extend(
        [
            "## Regeneration",
            "Run the generator again whenever the Excel workbook changes:",
            "",
            "```bash",
            "python3 scripts/generate_room_layout.py --input /path/to/숙소배정표.xlsx",
            "```",
            "",
            "The output files will be overwritten in `assets/`.",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def write_outputs(out_dir: Path, rooms: list[dict[str, Any]], spaces: list[dict[str, Any]], structure: dict[str, Any], workbook_path: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    room_layout_json = {
        "source": workbook_path.name,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "rooms": rooms,
        "service_spaces": spaces,
    }

    (out_dir / "room_layout.json").write_text(json.dumps(room_layout_json, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (out_dir / "building_structure.json").write_text(json.dumps(structure, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    csv_path = out_dir / "room_layout.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            lineterminator="\n",
            fieldnames=[
                "building",
                "floor",
                "room_number",
                "room_label",
                "room_type",
                "room_type_label",
                "capacity",
                "corridor_relationship",
                "cell",
                "row",
                "column",
                "source_text",
                "source_kind",
                "unavailable",
                "unavailable_reason",
            ],
        )
        writer.writeheader()
        for room in rooms:
            writer.writerow(
                {
                    "building": room["building"],
                    "floor": room["floor"],
                    "room_number": room["room_number"],
                    "room_label": room["room_label"],
                    "room_type": room["room_type"],
                    "room_type_label": room["room_type_label"],
                    "capacity": room["capacity"],
                    "corridor_relationship": room["corridor_relationship"],
                    "cell": room["cell"],
                    "row": room["row"],
                    "column": room["column"],
                    "source_text": room["source_text"],
                    "source_kind": room["source_kind"],
                    "unavailable": room.get("unavailable", False),
                    "unavailable_reason": room.get("unavailable_reason", ""),
                }
            )

    (out_dir / "room_type_rules.md").write_text(build_room_rules_md(structure, rooms), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate room layout assets from the retreat workbook.")
    parser.add_argument("--input", required=True, help="Path to 숙소배정표.xlsx")
    parser.add_argument("--output-dir", default="assets", help="Directory for generated assets")
    args = parser.parse_args()

    workbook_path = Path(args.input).expanduser().resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    sections = find_sections(ws)
    if not sections:
      raise SystemExit("No building/floor sections were detected in the workbook.")

    rooms, spaces = extract_room_entries(ws, sections)
    structure = group_structure(rooms, spaces, sections)
    write_outputs(Path(args.output_dir), rooms, spaces, structure, workbook_path)

    print(f"Generated {len(rooms)} rooms and {len(spaces)} service spaces from {workbook_path.name}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
