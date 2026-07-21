#!/usr/bin/env python3
"""Local static server for the retreat dashboard."""

from __future__ import annotations

import os
import json
import csv
import urllib.request
import io
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path


ROOT = Path(__file__).resolve().parent


class RetreatHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def do_GET(self):
        if self.path == '/api/family-db':
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            
            try:
                sheet_id = '1LDpA-cvYu9FS7e7eYSrxDPFHob2rOrJc_4jbpXi_-PE'
                csv_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv'
                
                req = urllib.request.Request(csv_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req) as response:
                    csv_data = response.read().decode('utf-8')
                
                reader = csv.DictReader(io.StringIO(csv_data))
                data = list(reader)
                
                self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                error_response = [{"error": str(e)}]
                self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
            return

        elif self.path == '/api/config':
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            try:
                load_env()
                config_data = {
                    "googleClientId": os.environ.get("GOOGLE_CLIENT_ID", "582144992351-iehddqkp7emjrkap1bh97p7fd5c3p8vf.apps.googleusercontent.com"),
                    "driveFolderId": os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "1WVtAhmSZ5OTZ9DOX0_afVPlegznir5KS")
                }
                self.wfile.write(json.dumps(config_data, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.wfile.write(json.dumps({"error": str(e)}, ensure_ascii=False).encode('utf-8'))
            return

        elif self.path.startswith('/api/drive-files'):
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            
            try:
                load_env()
                api_key = os.environ.get("GOOGLE_API_KEY", "")
                
                import urllib.parse
                parsed_path = urllib.parse.urlparse(self.path)
                params = urllib.parse.parse_qs(parsed_path.query)
                
                folder_id = params.get("folderId", ["1WVtAhmSZ5OTZ9DOX0_afVPlegznir5KS"])[0]
                
                # Load local uploads metadata
                local_uploads = []
                upload_dir = ROOT / 'uploaded_docs'
                meta_path = upload_dir / 'metadata.json'
                if meta_path.exists():
                    try:
                        with open(meta_path, 'r', encoding='utf-8') as f:
                            local_uploads = json.load(f)
                    except Exception:
                        pass
                
                local_uploads = [f for f in local_uploads if f.get("folderId") == folder_id]
                
                if not api_key:
                    # Return mock data + local uploads
                    response_data = {
                        "warning": "NO_API_KEY",
                        "files": get_mock_files_for_folder(folder_id) + local_uploads
                    }
                    self.wfile.write(json.dumps(response_data, ensure_ascii=False).encode('utf-8'))
                    return
                
                # Fetch files from Google Drive API
                q_query = f"'{folder_id}' in parents and trashed = false"
                fields = "files(id,name,mimeType,size,modifiedTime,webViewLink,owners)"
                drive_url = f"https://www.googleapis.com/drive/v3/files?q={urllib.parse.quote(q_query)}&fields={urllib.parse.quote(fields)}&key={api_key}"
                
                req = urllib.request.Request(drive_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req) as response:
                    res_data = json.loads(response.read().decode('utf-8'))
                
                files = res_data.get("files", []) + local_uploads
                self.wfile.write(json.dumps({"files": files}, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.wfile.write(json.dumps({"error": str(e)}, ensure_ascii=False).encode('utf-8'))
            return
            
        super().do_GET()

    def do_POST(self):
        if self.path == '/api/drive-upload':
            try:
                content_type = self.headers.get('Content-Type', '')
                content_length = int(self.headers.get('Content-Length', 0))
                
                body = self.rfile.read(content_length)
                
                boundary = content_type.split("boundary=")[-1].encode('utf-8')
                parts = body.split(boundary)
                
                file_data = b""
                filename = "uploaded_file"
                folder_id = "1WVtAhmSZ5OTZ9DOX0_afVPlegznir5KS"
                
                for part in parts:
                    if b'filename=' in part:
                        try:
                            headers_part, file_data_part = part.split(b'\r\n\r\n', 1)
                            if file_data_part.endswith(b'\r\n--'):
                                file_data_part = file_data_part[:-4]
                            elif file_data_part.endswith(b'\r\n'):
                                file_data_part = file_data_part[:-2]
                            
                            file_data = file_data_part
                            for line in headers_part.decode('utf-8', errors='ignore').split('\r\n'):
                                if 'filename=' in line:
                                    filename = line.split('filename=')[-1].strip('"')
                                    break
                        except Exception as e:
                            print("Error parsing part:", e)
                    elif b'name="folderId"' in part:
                        try:
                            _, val_part = part.split(b'\r\n\r\n', 1)
                            folder_id = val_part.strip().decode('utf-8')
                        except Exception:
                            pass
                
                # Save file locally
                upload_dir = ROOT / 'uploaded_docs'
                upload_dir.mkdir(exist_ok=True)
                
                save_path = upload_dir / filename
                with open(save_path, 'wb') as f:
                    f.write(file_data)
                
                # Save metadata
                meta_path = upload_dir / 'metadata.json'
                metadata = []
                if meta_path.exists():
                    try:
                        with open(meta_path, 'r', encoding='utf-8') as f:
                            metadata = json.load(f)
                    except Exception:
                        pass
                
                import datetime
                import os
                new_file_meta = {
                    "id": f"local_{os.urandom(4).hex()}",
                    "name": filename,
                    "mimeType": "application/octet-stream",
                    "size": str(len(file_data)),
                    "modifiedTime": datetime.datetime.now().isoformat() + "Z",
                    "webViewLink": f"/uploaded_docs/{filename}",
                    "folderId": folder_id,
                    "owners": [{"displayName": "로컬 사용자"}]
                }
                metadata.append(new_file_meta)
                with open(meta_path, 'w', encoding='utf-8') as f:
                    json.dump(metadata, f, ensure_ascii=False, indent=2)
                
                self.send_response(200)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                
                self.wfile.write(json.dumps({"success": True, "file": new_file_meta}, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"success": False, "message": str(e)}, ensure_ascii=False).encode('utf-8'))
            return
            
        elif self.path == '/api/export-room-reservation':
            try:
                import openpyxl
                import re
                
                content_length = int(self.headers.get('Content-Length', 0))
                body = self.rfile.read(content_length)
                req_data = json.loads(body.decode('utf-8'))
                
                assignments = req_data.get("assignments", {}) # { familyId: roomValue }
                families = req_data.get("families", [])       # list of family dicts
                
                # Create a map for quick family lookup
                family_map = {str(f.get("id")): f for f in families}
                
                # Group family display texts by normalized room name
                room_occupants = {}
                
                # Helpers to extract building & room number from roomValue string
                def parse_room_val(room_val):
                    if not room_val or room_val == "미배정":
                        return None
                    b_name = None
                    if "휴락동" in room_val:
                        b_name = "휴락동"
                    elif "동락홀" in room_val:
                        b_name = "동락홀"
                    
                    r_match = re.search(r"(\d{3})", room_val)
                    if r_match:
                        return (b_name, int(r_match.group(1)))
                    return None

                for f_id, room_val in assignments.items():
                    room_info = parse_room_val(room_val)
                    if not room_info:
                        continue
                    
                    fam = family_map.get(str(f_id))
                    if not fam:
                        continue
                    
                    fam_name = fam.get("name", "이름 없음")
                    nights_label = fam.get("nights_label", "")
                    
                    label = fam_name
                    if nights_label:
                        label += f" ({nights_label})"
                    
                    if room_info not in room_occupants:
                        room_occupants[room_info] = []
                    room_occupants[room_info].append(label)

                # Load original template workbook
                excel_path = ROOT / "Room_Reservation.xlsx"
                if not excel_path.exists():
                    raise FileNotFoundError(f"Template Room_Reservation.xlsx not found at {excel_path}")
                
                wb = openpyxl.load_workbook(excel_path)
                sheet = wb[wb.sheetnames[0]]
                
                # Scan sections to determine building/floor context for each cell
                SECTION_RE = re.compile(r".*?(?P<building>휴락동|동락홀|.+?)(?P<floor>\d)층")
                sections = []
                for row in range(1, sheet.max_row + 1):
                    val = sheet.cell(row, 2).value
                    if not val:
                        continue
                    normalized = re.sub(r"\s+", "", str(val))
                    match = SECTION_RE.match(normalized)
                    if match:
                        b_name = match.group("building").replace("▶", "").strip()
                        f_num = int(match.group("floor"))
                        sections.append((row, b_name, f_num))
                
                # Function to get building for a given row
                def get_building_for_row(row_idx):
                    current_b = None
                    for r, b, f in sections:
                        if row_idx >= r:
                            current_b = b
                        else:
                            break
                    return current_b

                # Helper to check if a cell is in a merged range and return top-left
                def get_merged_top_left(r, c):
                    for m_range in sheet.merged_cells.ranges:
                        if r >= m_range.min_row and r <= m_range.max_row and c >= m_range.min_col and c <= m_range.max_col:
                            return m_range.min_row, m_range.min_col
                    return None

                ROOM_RE = re.compile(r"^(\d{3})호")
                processed_merged_cells = set()

                for row in range(1, sheet.max_row + 1):
                    building = get_building_for_row(row)
                    if not building:
                        continue
                        
                    for col in range(1, sheet.max_column + 1):
                        val = sheet.cell(row, col).value
                        if not val:
                            continue
                        
                        normalized_val = re.sub(r"\s+", "", str(val))
                        room_match = ROOM_RE.match(normalized_val)
                        if room_match:
                            room_number = int(room_match.group(1))
                            key = (building, room_number)
                            
                            if key in room_occupants:
                                occupants_text = "\n".join(room_occupants[key])
                                
                                target_row = row + 1
                                target_merged = get_merged_top_left(target_row, col)
                                if target_merged:
                                    t_r, t_c = target_merged
                                    if (t_r, t_c) not in processed_merged_cells:
                                        orig_val = sheet.cell(t_r, t_c).value or ""
                                        separator = "\n" if orig_val else ""
                                        sheet.cell(t_r, t_c).value = f"{orig_val}{separator}{occupants_text}"
                                        processed_merged_cells.add((t_r, t_c))
                                else:
                                    sheet.cell(target_row, col).value = occupants_text

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename=Room_Reservation_assigned.xlsx')
                self.end_headers()
                self.wfile.write(output.read())
                return
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"success": False, "message": str(e)}, ensure_ascii=False).encode('utf-8'))
                return
            
        self.send_response(404)
        self.end_headers()

def load_env():
    env_path = ROOT / '.env'
    if env_path.exists():
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, val = line.split('=', 1)
                    os.environ[key.strip()] = val.strip()

def get_mock_files_for_folder(folder_id):
    if folder_id == "1WVtAhmSZ5OTZ9DOX0_afVPlegznir5KS":
        return [
            { "id": "folder_admin", "name": "기획 및 행정", "mimeType": "application/vnd.google-apps.folder", "owners": [{"displayName": "시스템"}] },
            { "id": "folder_budget", "name": "예산 및 재정", "mimeType": "application/vnd.google-apps.folder", "owners": [{"displayName": "시스템"}] },
            { "id": "folder_program", "name": "프로그램 & 악보", "mimeType": "application/vnd.google-apps.folder", "owners": [{"displayName": "시스템"}] },
            { "id": "folder_guide", "name": "안내 및 매뉴얼", "mimeType": "application/vnd.google-apps.folder", "owners": [{"displayName": "시스템"}] }
        ]
    
    mock_db = [
      { "name": "2026_여름수련회_기획안.pdf", "mimeType": "application/pdf", "size": "2516582", "folder": "기획 및 행정", "modifiedTime": "2026-06-01T10:00:00.000Z", "owners": [{"displayName": "조봄이와"}], "webViewLink": "https://docs.google.com/document/d/1BxiMVs0XRA5nFMdKv3dBt28vJUX078y1/edit?usp=sharing" },
      { "name": "수련회_예산안_및_집행계획.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "size": "1887436", "folder": "예산 및 재정", "modifiedTime": "2026-05-28T14:30:00.000Z", "owners": [{"displayName": "김상배"}], "webViewLink": "https://docs.google.com/spreadsheets/d/1BxiMVs0XRA5nFMdKv3dBt28vJUX078y1/edit?usp=sharing" },
      { "name": "숙소배정표_최종본.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "size": "972800", "folder": "기획 및 행정", "modifiedTime": "2026-06-05T09:15:00.000Z", "owners": [{"displayName": "윤선욱"}], "webViewLink": "https://docs.google.com/spreadsheets/d/1BxiMVs0XRA5nFMdKv3dBt28vJUX078y1/edit?usp=sharing" },
      { "name": "식사수요조사_결과보고.docx", "mimeType": "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "size": "1258291", "folder": "안내 및 매뉴얼", "modifiedTime": "2026-06-02T11:20:00.000Z", "owners": [{"displayName": "김미현"}], "webViewLink": "https://docs.google.com/document/d/1BxiMVs0XRA5nFMdKv3dBt28vJUX078y1/edit?usp=sharing" },
      { "name": "수련회_안내문_및_준비물.pdf", "mimeType": "application/pdf", "size": "3250585", "folder": "안내 및 매뉴얼", "modifiedTime": "2026-06-04T16:00:00.000Z", "owners": [{"displayName": "장세연"}], "webViewLink": "https://docs.google.com/document/d/1BxiMVs0XRA5nFMdKv3dBt28vJUX078y1/edit?usp=sharing" },
      { "name": "찬양콘서트_콘티_및_악보.pdf", "mimeType": "application/pdf", "size": "8912896", "folder": "프로그램 & 악보", "modifiedTime": "2026-06-03T18:45:00.000Z", "owners": [{"displayName": "이지은A"}], "webViewLink": "https://docs.google.com/document/d/1BxiMVs0XRA5nFMdKv3dBt28vJUX078y1/edit?usp=sharing" },
      { "name": "안전관리_및_비상연락망.pdf", "mimeType": "application/pdf", "size": "870400", "folder": "안내 및 매뉴얼", "modifiedTime": "2026-05-30T10:10:00.000Z", "owners": [{"displayName": "박철진"}], "webViewLink": "https://docs.google.com/document/d/1BxiMVs0XRA5nFMdKv3dBt28vJUX078y1/edit?usp=sharing" },
      { "name": "조별성경공부_가이드북.pdf", "mimeType": "application/pdf", "size": "4404019", "folder": "프로그램 & 악보", "modifiedTime": "2026-06-02T15:30:00.000Z", "owners": [{"displayName": "박철진"}], "webViewLink": "https://docs.google.com/document/d/1BxiMVs0XRA5nFMdKv3dBt28vJUX078y1/edit?usp=sharing" }
    ]
    
    folder_name_map = {
      "folder_admin": "기획 및 행정",
      "folder_budget": "예산 및 재정",
      "folder_program": "프로그램 & 악보",
      "folder_guide": "안내 및 매뉴얼"
    }
    target_folder = folder_name_map.get(folder_id, folder_id)
    return [f for f in mock_db if f.get("folder") == target_folder]



if __name__ == "__main__":
    port = int(os.environ.get("PORT", "4173"))
    server = ThreadingHTTPServer(("127.0.0.1", port), RetreatHandler)
    print(f"Retreat dashboard server running at http://127.0.0.1:{port}/")
    server.serve_forever()
